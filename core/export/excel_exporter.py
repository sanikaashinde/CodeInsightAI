from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelExporter:

    def generate(
        self,
        project_name,
        metrics,
        quality,
        security_score,
        libraries,
        duplicate_summary,
        todo_summary,
    ):

        wb = Workbook()

        # ==========================================
        # Overview Sheet
        # ==========================================

        ws = wb.active
        ws.title = "Overview"

        ws["A1"] = "CodeInsight AI Report"
        ws["A1"].font = Font(bold=True, size=16)

        ws["A3"] = "Project"
        ws["B3"] = project_name

        row = 5

        for key, value in metrics.items():

            ws.cell(row=row, column=1).value = key.replace("_", " ").title()
            ws.cell(row=row, column=2).value = value

            row += 1

        row += 1

        for key, value in quality.items():

            ws.cell(row=row, column=1).value = key.title()
            ws.cell(row=row, column=2).value = value

            row += 1

        row += 1

        ws.cell(row=row, column=1).value = "Security Score"
        ws.cell(row=row, column=2).value = security_score

        # ==========================================
        # Libraries
        # ==========================================

        sheet = wb.create_sheet("Libraries")

        sheet.append(

            [

                "Library",

                "Usage"

            ]

        )

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for lib, count in libraries:

            sheet.append(

                [

                    lib,

                    count

                ]

            )

        # ==========================================
        # Duplicate Code
        # ==========================================

        dup = wb.create_sheet("Duplicate Code")

        dup.append(

            [

                "Metric",

                "Value"

            ]

        )

        dup["A1"].font = Font(bold=True)
        dup["B1"].font = Font(bold=True)

        for k, v in duplicate_summary.items():

            dup.append(

                [

                    k,

                    v

                ]

            )

        # ==========================================
        # TODO Summary
        # ==========================================

        todo = wb.create_sheet("TODOs")

        todo.append(

            [

                "Type",

                "Count"

            ]

        )

        todo["A1"].font = Font(bold=True)
        todo["B1"].font = Font(bold=True)

        for k, v in todo_summary.items():

            todo.append(

                [

                    k,

                    v

                ]

            )

        # ==========================================
        # Export
        # ==========================================

        output = BytesIO()

        wb.save(output)

        output.seek(0)

        return output.getvalue()